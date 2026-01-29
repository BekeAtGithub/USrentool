import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def deep_scan_excel(filepath):
    """Perform a comprehensive deep scan of the Excel file."""
    print(f"\n{'='*100}")
    print(f"DEEP ANALYSIS: {filepath}")
    print(f"{'='*100}\n")
    
    try:
        # Load workbook with formulas
        wb = openpyxl.load_workbook(filepath, data_only=False)
        
        for sheet_name in wb.sheetnames:
            print(f"\n{'*'*100}")
            print(f"SHEET: {sheet_name}")
            print(f"{'*'*100}\n")
            
            sheet = wb[sheet_name]
            
            # Get all merged cells
            print("MERGED CELLS:")
            for merged_range in sheet.merged_cells.ranges:
                print(f"  {merged_range}")
            
            print(f"\nDIMENSIONS: {sheet.max_row} rows x {sheet.max_column} columns")
            
            # Scan all cells for content, formulas, and formatting
            print("\nDETAILED CELL CONTENT (First 100 rows):\n")
            
            for row_idx in range(1, min(101, sheet.max_row + 1)):
                row_data = []
                has_content = False
                
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    if cell.value is not None:
                        has_content = True
                        col_letter = get_column_letter(col_idx)
                        
                        # Get cell details
                        value = cell.value
                        formula = None
                        if isinstance(value, str) and value.startswith('='):
                            formula = value
                            # Try to get calculated value
                            wb_data = openpyxl.load_workbook(filepath, data_only=True)
                            calculated_value = wb_data[sheet_name].cell(row=row_idx, column=col_idx).value
                            wb_data.close()
                            value = calculated_value
                        
                        # Check for fill color
                        fill_color = None
                        if cell.fill and cell.fill.start_color:
                            fill_color = cell.fill.start_color.rgb
                        
                        # Check for comments
                        comment = cell.comment.text if cell.comment else None
                        
                        cell_info = {
                            'cell': f'{col_letter}{row_idx}',
                            'value': value,
                            'formula': formula,
                            'fill': fill_color,
                            'comment': comment
                        }
                        row_data.append(cell_info)
                
                if has_content:
                    print(f"\nRow {row_idx}:")
                    for cell_info in row_data:
                        print(f"  [{cell_info['cell']}] = {cell_info['value']}")
                        if cell_info['formula']:
                            print(f"    Formula: {cell_info['formula']}")
                        if cell_info['fill'] and cell_info['fill'] != '00000000':
                            print(f"    Fill Color: {cell_info['fill']}")
                        if cell_info['comment']:
                            print(f"    Comment: {cell_info['comment']}")
            
            # Now try to read with pandas to get structured data
            print(f"\n\n{'='*100}")
            print(f"PANDAS DATAFRAME VIEW (First 50 rows):")
            print(f"{'='*100}\n")
            
            try:
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
                print(df.head(50).to_string())
                
                # Look for patterns
                print(f"\n\nDATA PATTERNS:")
                print(f"Total cells with data: {df.notna().sum().sum()}")
                print(f"Columns with data: {df.notna().any().sum()}")
                print(f"Rows with data: {df.notna().any(axis=1).sum()}")
                
            except Exception as e:
                print(f"Could not read with pandas: {e}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error analyzing file: {e}")
        import traceback
        traceback.print_exc()

def main():
    file = "Deal Analysis Cheat Sheet.xlsx"
    deep_scan_excel(file)
    
    print(f"\n\n{'='*100}")
    print("DEEP ANALYSIS COMPLETE")
    print(f"{'='*100}")

if __name__ == "__main__":
    main()

