import pandas as pd
import openpyxl
import os
from pathlib import Path

def analyze_excel_file(filepath):
    """Analyze an Excel file and extract information about its structure and content."""
    print(f"\n{'='*80}")
    print(f"ANALYZING: {filepath}")
    print(f"{'='*80}\n")
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        print(f"Number of sheets: {len(wb.sheetnames)}")
        print(f"Sheet names: {', '.join(wb.sheetnames)}\n")
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            sheet = wb[sheet_name]
            
            # Get dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"Dimensions: {max_row} rows x {max_col} columns")
            
            # Try to read with pandas for better data display
            try:
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
                print(f"\nFirst 20 rows of data:")
                print(df.head(20).to_string())
                
                # Show some statistics if there are numeric columns
                numeric_df = df.select_dtypes(include=['number'])
                if not numeric_df.empty:
                    print(f"\nNumeric columns found: {len(numeric_df.columns)}")
                    
            except Exception as e:
                print(f"Could not read sheet with pandas: {e}")
                
                # Fallback to openpyxl
                print("\nSample data (first 20 rows):")
                for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if i <= 20:
                        print(f"Row {i}: {row}")
                    else:
                        break
        
        wb.close()
        
    except Exception as e:
        print(f"Error analyzing file: {e}")

def main():
    # Get all Excel files in current directory
    excel_files = [
        "Deal Analysis Cheat Sheet.xlsx",
        "InvestwithACE_Property_Management_Interview_Guide.xlsx",
        "InvestwithAce The Legendary Rental Property Calculator.xlsx",
        "Lending Starter Kit + Net Worth Manager.xlsx"
    ]
    
    for file in excel_files:
        if os.path.exists(file):
            analyze_excel_file(file)
        else:
            print(f"\nFile not found: {file}")
    
    print(f"\n{'='*80}")
    print("ANALYSIS COMPLETE")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()

